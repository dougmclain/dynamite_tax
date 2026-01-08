# Update tax_form/views/export.py

from django.views import View
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.utils import timezone
from ..models import Association, Financial, CompletedTaxReturn, ManagementCompany, AssociationFilingStatus
import csv
import logging
from io import BytesIO

logger = logging.getLogger(__name__)

class ExportAssociationsView(LoginRequiredMixin, View):
    """View to export association data for a specific management company, filtered by filing status"""
    
    def get(self, request):
        # Get parameters from the request
        management_company_id = request.GET.get('management_company')
        tax_year = request.GET.get('tax_year')
        export_format = request.GET.get('format', 'html')  # Default to HTML
        
        if not tax_year:
            # Default to current year
            tax_year = timezone.now().year
        else:
            tax_year = int(tax_year)
        
        # Check if a specific management company is selected
        if not management_company_id or management_company_id == 'all':
            # Redirect back to dashboard with an error message if no specific company is selected
            messages.error(request, "Please select a specific management company before exporting.")
            return redirect(f'/dashboard/?tax_year={tax_year}&management_company=all')
        
        # Define the query for associations
        query = Association.objects.all().order_by('association_name')
        management_company_name = "All Companies"
        
        # Filter by management company
        if management_company_id == 'self':
            query = query.filter(is_self_managed=True)
            management_company_name = "Self-Managed"
        else:
            # Filter for specific management company
            try:
                company = ManagementCompany.objects.get(id=management_company_id)
                query = query.filter(management_company=company)
                management_company_name = company.name
            except ManagementCompany.DoesNotExist:
                messages.error(request, "Selected management company not found.")
                return redirect(f'/dashboard/?tax_year={tax_year}')
        
        # Get associations with their tax return status
        associations_data = []
        
        # Counter for numbering only included associations
        idx = 1
        
        for association in query:
            # Check if this association is marked for filing in the selected tax year
            filing_status, created = AssociationFilingStatus.objects.get_or_create(
                association=association,
                tax_year=tax_year,
                defaults={'prepare_return': True}
            )
            
            # Skip associations not marked for filing
            if not filing_status.prepare_return:
                continue
            
            # Get financial record for this tax year
            financial = Financial.objects.filter(
                association=association,
                tax_year=tax_year
            ).first()
            
            # Get tax return status
            completed_tax_return = None
            sent_status = "Not Sent"
            filed_status = "Not Filed"
            
            if financial:
                completed_tax_return = CompletedTaxReturn.objects.filter(
                    financial=financial
                ).first()
                
                if completed_tax_return:
                    if completed_tax_return.sent_for_signature:
                        sent_date = ""
                        if completed_tax_return.sent_date:
                            sent_date = f" ({completed_tax_return.sent_date.strftime('%m/%d/%Y')})"
                        sent_status = f"Sent{sent_date}"
                    
                    if completed_tax_return.return_filed:
                        filed_date = ""
                        if completed_tax_return.date_prepared:
                            filed_date = f" ({completed_tax_return.date_prepared.strftime('%m/%d/%Y')})"
                        
                        status_map = {
                            'not_filed': 'Not Filed',
                            'filed_by_dynamite': 'Filed by Dynamite',
                            'filed_by_association': 'Filed by Association'
                        }
                        filed_status = f"{status_map.get(completed_tax_return.filing_status, 'Filed')}{filed_date}"
            
            # Determine management info
            if association.is_self_managed:
                management_info = "Self-Managed"
            elif association.management_company:
                management_info = association.management_company.name
            else:
                management_info = "Unspecified"
            
            # Add to our data list with the current index
            associations_data.append({
                'number': idx,
                'name': association.association_name,
                'ein': association.ein,
                'management': management_info,
                'sent_status': sent_status,
                'filed_status': filed_status,
                'invoiced': 'Yes' if filing_status.invoiced else 'No'
            })
            
            # Increment the index only for associations we're including
            idx += 1
        
        # If no associations match our criteria, show a message
        if not associations_data:
            messages.warning(request, f"No associations marked for filing found for {management_company_name} in {tax_year}.")
            return redirect(f'/dashboard/?tax_year={tax_year}&management_company={management_company_id}')
        
        # Handle different export formats
        if export_format == 'csv':
            # Create a CSV response
            filename = f"associations_to_file_{management_company_name.replace(' ', '_')}_{tax_year}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Create CSV writer directly using response
            writer = csv.writer(response)
            
            # Write header
            writer.writerow(['#', 'Association Name', 'EIN', 'Management', 'Tax Year', 'Return Sent Status', 'Return Filed Status', 'Invoiced'])
            
            # Write data
            for assoc in associations_data:
                writer.writerow([
                    assoc['number'],
                    assoc['name'],
                    assoc['ein'],
                    assoc['management'],
                    tax_year,
                    assoc['sent_status'],
                    assoc['filed_status'],
                    assoc['invoiced']
                ])
            
            return response
        else:
            # Return HTML view
            context = {
                'associations_data': associations_data,
                'tax_year': tax_year,
                'management_company_name': management_company_name,
                'filing_only': True
            }
            return render(request, 'tax_form/export_associations.html', context)


class ExportCompletedReturnsExcelView(LoginRequiredMixin, View):
    """Export completed returns for a management company to Excel with a next year todo column"""

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        # Get parameters
        management_company_id = request.GET.get('management_company')
        source_year = request.GET.get('source_year')
        target_year = request.GET.get('target_year')

        if not source_year:
            source_year = timezone.now().year - 1  # Default to previous year
        else:
            source_year = int(source_year)

        if not target_year:
            target_year = source_year + 1
        else:
            target_year = int(target_year)

        # Build query for associations
        query = Association.objects.all().order_by('association_name')
        management_company_name = "All Companies"

        # Filter by management company
        if management_company_id == 'self':
            query = query.filter(is_self_managed=True)
            management_company_name = "Self-Managed"
        elif management_company_id and management_company_id != 'all':
            try:
                company = ManagementCompany.objects.get(id=management_company_id)
                query = query.filter(management_company=company)
                management_company_name = company.name
            except ManagementCompany.DoesNotExist:
                messages.error(request, "Selected management company not found.")
                return redirect('dashboard')

        # Collect associations with completed returns for source year
        associations_data = []

        for association in query:
            # Get financial record for source year
            financial = Financial.objects.filter(
                association=association,
                tax_year=source_year
            ).first()

            if not financial:
                continue

            # Get completed tax return
            completed_return = CompletedTaxReturn.objects.filter(
                financial=financial
            ).first()

            # Only include associations where the return was filed
            if not completed_return or not completed_return.return_filed:
                continue

            # Get filing status for source year
            filing_status = AssociationFilingStatus.objects.filter(
                association=association,
                tax_year=source_year
            ).first()

            # Determine sent date
            sent_date = ""
            if completed_return.sent_for_signature and completed_return.sent_date:
                sent_date = completed_return.sent_date.strftime('%m/%d/%Y')

            # Determine filed date
            filed_date = ""
            if completed_return.date_prepared:
                filed_date = completed_return.date_prepared.strftime('%m/%d/%Y')

            associations_data.append({
                'name': association.association_name,
                'ein': association.ein or '',
                'state': association.get_filing_state() if hasattr(association, 'get_filing_state') else association.state,
                'sent_date': sent_date,
                'filed_date': filed_date,
                'filing_status': completed_return.get_filing_status_display() if completed_return.filing_status else '',
                'invoiced': 'Yes' if filing_status and filing_status.invoiced else 'No',
            })

        if not associations_data:
            messages.warning(request, f"No completed returns found for {management_company_name} in {source_year}.")
            return redirect(f'/dashboard/?tax_year={source_year}&management_company={management_company_id or "all"}')

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{source_year} Completed Returns"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        checkbox_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        # Define headers
        headers = [
            '#',
            'Association Name',
            'EIN',
            'Filing State',
            f'{source_year} Sent Date',
            f'{source_year} Filed Date',
            'Filing Status',
            'Invoiced',
            f'{target_year} Return Todo'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        for row_idx, assoc in enumerate(associations_data, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=2, value=assoc['name']).border = thin_border
            ws.cell(row=row_idx, column=3, value=assoc['ein']).border = thin_border
            ws.cell(row=row_idx, column=4, value=assoc['state']).border = thin_border
            ws.cell(row=row_idx, column=5, value=assoc['sent_date']).border = thin_border
            ws.cell(row=row_idx, column=6, value=assoc['filed_date']).border = thin_border
            ws.cell(row=row_idx, column=7, value=assoc['filing_status']).border = thin_border
            ws.cell(row=row_idx, column=8, value=assoc['invoiced']).border = thin_border

            # Todo column - empty with yellow highlight for easy checking
            todo_cell = ws.cell(row=row_idx, column=9, value="")
            todo_cell.border = thin_border
            todo_cell.fill = checkbox_fill
            todo_cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        column_widths = [5, 45, 15, 12, 15, 15, 18, 10, 18]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add summary row at the top as a second sheet
        summary_ws = wb.create_sheet(title="Summary")
        summary_ws.cell(row=1, column=1, value="Export Summary").font = Font(bold=True, size=14)
        summary_ws.cell(row=3, column=1, value="Management Company:")
        summary_ws.cell(row=3, column=2, value=management_company_name)
        summary_ws.cell(row=4, column=1, value="Source Year:")
        summary_ws.cell(row=4, column=2, value=source_year)
        summary_ws.cell(row=5, column=1, value="Target Year:")
        summary_ws.cell(row=5, column=2, value=target_year)
        summary_ws.cell(row=6, column=1, value="Total Associations:")
        summary_ws.cell(row=6, column=2, value=len(associations_data))
        summary_ws.cell(row=7, column=1, value="Export Date:")
        summary_ws.cell(row=7, column=2, value=timezone.now().strftime('%m/%d/%Y %H:%M'))

        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 30

        # Generate filename
        safe_company_name = management_company_name.replace(' ', '_').replace('/', '-')
        filename = f"completed_returns_{safe_company_name}_{source_year}_with_{target_year}_todo.xlsx"

        # Write to response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response